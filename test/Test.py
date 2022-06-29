
from tkinter import *
from openpyxl import *
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo

root = Tk()
root.geometry("900x600+250+100")

root.title("AbOt")
root.configure(bg='#464648') # to`q kulrang


baza = []
def cheak():
    global m, k, status
    login = e1.get()
    password = e2.get()
    status = 0
    k = 0
    m = 0
    
    
    book = load_workbook('Baza.xlsx')
    sheet = book.active
    for i in range(1, sheet.max_row+1):
        A = f'A{i}'
        B = f'B{i}'
        C = f'C{i}'
        D = f'D{i}'
        E = f'E{i}'
      
        baza.append([sheet[A].value, sheet[B].value, sheet[C].value, sheet[D].value, sheet[E].value])
    
    for i in baza:
        if login == i[3] and password == str(i[4]):
            status += 1
            m = k

        k += 1
        

    
    if status == 1:
        
        l5 = Label(frame_1, text = f"Hurmatli {baza[m][0]} {baza[m][1]} tizimga hush kelibsiz", font=("Arial", 15), bg="#CCCDD0", fg='#464648')
        l5.place(y = 210, x = 100, height = 30, width = 600)
        boshlash()
    else:
        
        l5 = Label(frame_1, text = f"ERROR", font=("Arial", 15), bg="#CCCDD0", fg='#464648')
        l5.place(y = 210, x = 100, height = 30, width = 600)
        main()
    

def add_baza():
    baza_2.append(e6.get())
    baza_2.append(e7.get())

    wb = Workbook()
    ws = wb.active
    
    book = load_workbook('Baza.xlsx')
    sheet = book.active
    for i in range(1, sheet.max_row+1):
        A = f'A{i}'
        B = f'B{i}'
        C = f'C{i}'
        D = f'D{i}'
        E = f'E{i}'
        F = f'F{i}'
        G = f'G{i}'
        H = f'H{i}'
        I = f'I{i}'
        J = f'J{i}'
        K = f'K{i}'
        baza.append([sheet[A].value, sheet[B].value, sheet[C].value, sheet[D].value, sheet[E].value, sheet[F].value, sheet[G].value, sheet[H].value, sheet[I].value, sheet[J].value, sheet[K].value])
    baza.append(baza_2)
    for i in baza:
        ws.append(i)
    wb.save("Baza.xlsx")
    main()
def main():
    global frame_1,e1,e2
    
    frame_1 = Frame(root, bg='#CCCDD0')
    frame_1.place(x = 100, y = 100, width=700, height=400)


    l1 = Label(frame_1, text="Dasturimizga hush kelibsiz", font=("Arial", 15), bg="#CCCDD0", fg='#464648')
    l1.place(x = 200, y = 10, height=30, width=300)




    l2 = Label(frame_1, text="Log In", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l2.place(x = 100, y = 100, height=30, width=80)


    l3 = Label(frame_1, text="Password", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l3.place(x = 100, y = 150, height=30, width=80)


    e1 = Entry(frame_1)
    e1.place(x = 200, y = 100, height = 30, width = 200)


    e2 = Entry(frame_1, show = "*")
    e2.place(x = 200, y = 150, height = 30, width = 200)
    def close():
        root.destroy()
        

    btn_1 = Button(frame_1, text = "Exit", bd = 6, activebackground = 'black', activeforeground = "white", command = close)
    btn_1.place(x = 500, y = 350, height = 30, width = 80)


    btn_2 = Button(frame_1, text = "Ok", bd = 6, activebackground = 'black', activeforeground = "white", command = cheak)
    btn_2.place(x = 600, y = 350, height = 30, width = 80)

    btn_3 = Button(frame_1, text = "Ro`yxatdan o`tish", bd = 6, activebackground = 'black', activeforeground = "white", command = ruyxat)
    btn_3.place(x = 50, y = 350, height = 30, width = 120)
    
    l5 = Label(frame_1, text ="", bg="#CCCDD0", fg='#464648')
    l5.place(y = 210, x = 100, height = 30, width = 600)

    
def ruyxat():
    global frame_2, e3, e4, e5, baza_2

    baza_2 = []
    frame_1.destroy()
    
    
    frame_2 = Frame(root, bg='#CCCDD0') #och kulrang
    frame_2.place(x = 100, y = 100, width=700, height=400)
    

    l1 = Label(frame_2, text="Ruyxatdan o`tish", font=("Arial", 15), bg="#CCCDD0", fg='#464648')
    l1.place(x = 200, y = 15, height=30, width=300)

    l2 = Label(frame_2, text="Ismingizni kiriting:", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l2.place(x = 100, y = 100, height=30, width=150)

    e3 = Entry(frame_2)
    e3.place(x = 270, y = 100, height = 30, width = 200)
    
    l3 = Label(frame_2, text="Familiyangizni kiriting:", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l3.place(x = 100, y = 150, height=30, width=150)
    
    e4 = Entry(frame_2)
    e4.place(x = 270, y = 150, height = 30, width = 200)

    l4 = Label(frame_2, text="Tug`ilgan sana :", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l4.place(x = 100, y = 200, height=30, width=150)

    e5 = Entry(frame_2)
    e5.place(x = 270, y = 200, height = 30, width = 200)

    btn_5 = Button(frame_2, text = "Next", bd = 6, activebackground = 'black', activeforeground = "white", command = login)
    btn_5.place(x = 600, y = 350, height = 30, width = 80)

    
def login():
    global frame_3, e6, e7

    baza_2.append(e3.get())
    baza_2.append(e4.get())
    baza_2.append(e5.get())

    frame_2.destroy()
    

    frame_3 = Frame(root, bg='#CCCDD0')
    frame_3.place(x = 100, y = 100, width=700, height=400)

    l1 = Label(frame_3, text="Login password", font=("Arial", 15), bg="#CCCDD0", fg='#464648')
    l1.place(x = 200, y = 10, height=30, width=300)

    l2 = Label(frame_3, text="Yangi login kiriting:", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l2.place(x = 100, y = 100, height=30, width=150)

    e6 = Entry(frame_3)
    e6.place(x = 270, y = 100, height = 30, width = 200)

    l3 = Label(frame_3, text="Password:", font=("Arial", 10), bg="#CCCDD0", fg='#464648')
    l3.place(x = 100, y = 150, height=30, width=150)

    e7 = Entry(frame_3, show = "*")
    e7.place(x = 270, y = 150, height = 30, width = 200)

    btn_6 = Button(frame_3, text = "Ok", bd = 6, activebackground = 'black', activeforeground = "white", command = add_baza)
    btn_6.place(x = 600, y = 350, height = 30, width = 80)

def select_file():
    global filename
    filetypes = (
        ('xlsx files', '*.xlsx'),
        ('All files', '*.*')
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)
    test()

def test():
    global d,s,minut,frame_4_1,t
    
    
    frame_4_1 = Frame(root, bg="#4572F9")
    frame_4_1.place(x = 0, y = 0, height=50, width=900)
    
    

    l2 = Label(frame_4_1, text = f"{baza[m][0]} {baza[m][1]}", font=("Arial", 15), bg="#4572F9", fg='white')
    l2.place(x = 300, y = 15, height=30, width=300)

    

    frame_4_2 = Frame(root, bg="#3B4B7D")
    frame_4_2.place(x = 0, y = 550, height=50, width=900)


    frame_4_3 = Frame(root, bg="#1B2A58")
    frame_4_3.place(x = 0, y = 50, height=500, width=900)
    try:
        book = load_workbook(filename)
        sheet = book.active
        savol = []
        Ak = []
        Bk = []
        Ck = []
        Dk = []
        tj = []
        for i in range(1, sheet.max_row+1):
            A = f'A{i}'
            B = f'B{i}'
            C = f'C{i}'
            D = f'D{i}'
            E = f'E{i}'
            F = f'F{i}'
      
            #baza.append([sheet[A].value, sheet[B].value, sheet[C].value, sheet[D].value, sheet[E].value, sheet[F].value])
            savol.append(sheet[A].value)
            Ak.append(sheet[B].value)
            Bk.append(sheet[C].value)
            Ck.append(sheet[D].value)
            Dk.append(sheet[E].value)
            tj.append(sheet[F].value)
        '''for i in baza:
        print(i)
        print(type(password))
        if login == i[3] and password == str(i[4]):
            status += 1
            m = k

        k += 1'''
    
        t = (sheet.max_row) * 1.5
        minut = int(t // 1)
        s = int((t % 1)*60)
    
        
        def update():
            global minut, s,frame_4_1
            if m < 10 and s < 10:
                l3 = Label(frame_4_1, text=f"Qolgan vaqt: 0{minut} : 0{s}", font=("Arial", 15), bg="#4572F9", fg='white')
                l3.place(x = 600, y = 15, height=30, width=300)   
                s -= 1    
                l3.after(1000,update)
        
                if s== -1 and minut == 0:
                    natija()
                elif s < 0:
                    minut -=1
                    s = 60
            elif s< 10:
                l3 = Label(frame_4_1, text=f"Qolgan vaqt: {minut} : 0{s}", font=("Arial", 15), bg="#4572F9", fg='white')
                l3.place(x = 600, y = 15, height=30, width=300)  
                s -= 1    
                l3.after(1000,update)
        
                if s== -1 and minut == 0:
                    natija()
                elif s < 0:
                    minut -=1
                    s = 60
            elif minut < 10:
                l3 = Label(frame_4_1, text=f"Qolgan vaqt: 0{minut} : {s}", font=("Arial", 15), bg="#4572F9", fg='white')
                l3.place(x = 600, y = 15, height=30, width=300)   
                s -= 1    
                l3.after(1000,update)
        
                if s== -1 and minut == 0:
                    natija()
                elif s < 0:
                    minut -=1
                    s = 60
            else:
                l3 = Label(frame_4_1, text=f"Qolgan vaqt: 0{minut} : 0{s}", font=("Arial", 15), bg="#4572F9", fg='white')
                l3.place(x = 600, y = 15, height=30, width=300)    
                s -= 1    
                l3.after(1000,update)
        
                if s== -1 and minut == 0:
                    natija()
                elif s < 0:
                    minut -=1
                    s = 60
        
        update()
        
        global sanoq

        sanoq = -1
        def oldinga():
            l3_1 = Label(frame_4_3, text="", font=("Arial", 15), bg="#1B2A58", fg='white')
            l3_1.place(x = 150, y = 200, height=30, width=600)
            global sanoq
            if sanoq == sheet.max_row-2:
                finish()

            sanoq += 1
            test_prog(savol[sanoq],Ak[sanoq] , Bk[sanoq], Ck[sanoq], Dk[sanoq])
            
        d ={}    
        for i in range(sheet.max_row):
            d[i] = 0        
        def finish():
            
            btn_8_1 = Button(frame_4_3, text = "Finish", bd = 6, activebackground = '#1C4BD6', activeforeground = "white", fg='white', bg = '#001659', command = natija )
            btn_8_1.place(x = 800, y = 450, height = 30, width = 80)
        def natija():
            frame_4_1.destroy()
            frame_4_2.destroy()
            frame_4_3.destroy()

            frame_5 = Frame(root, bg="#4572F9")
            frame_5.place(x = 0, y = 0, height=50, width=900)
        
        

            l2 = Label(frame_5, text = f"{baza[m][0]} {baza[m][1]}", font=("Arial", 15), bg="#4572F9", fg='white')
            l2.place(x = 300, y = 15, height=30, width=300)

            frame_4_4 = Frame(root, bg='#1B2A58')
            frame_4_4.place(x = 0, y = 50, height=500, width=900)
            
            result = 0
            for i in range(sheet.max_row):
                if d[i] == tj[i]:
                    result += 1
                    
            l3_1_2 = Label(frame_4_4, text=f"Testning natijasi: {100*(result/sheet.max_row) } %", font=("Arial", 15), bg="#4572F9", fg='white')
            l3_1_2.place(x = 150, y = 200, height=30, width=600)

        

        def ortga():
            global sanoq,d
            if sanoq == 0:
                test_prog(savol[0],Ak[0] , Bk[0], Ck[0], Dk[0])
                sanoq = 1
            sanoq -= 1
            test_prog(savol[sanoq],Ak[sanoq] , Bk[sanoq], Ck[sanoq], Dk[sanoq])

        
    
        def test_prog(savol, A_key, B_key, C_key, D_key):
            l11 = Label(frame_4_1, text=f"Test {sanoq + 1} ", font=("Arial", 15), bg="#4572F9", fg='white')
            l11.place(x = 0, y = 15, height=30, width=300)
            
            global d


            l1 = Label(frame_4_3, text=savol, font=("Arial", 15), bg="white", fg="#4572F9")
            l1.place(x = 20, y = 20, height = 50, width = 860)
            s = 0
            def sellect():
                global d
                d[sanoq] = radio.get()
        
            radio = IntVar()  
            radio.set(d[sanoq])
            
            R1 = Radiobutton(frame_4_3, text=f"A) {A_key}",  variable=radio, value=1,fg="#4572F9", command = sellect)  
            R1.place(x = 20, y=120, height=50, width=860 , anchor = "w")
            
        
            R2 = Radiobutton(frame_4_3, text=f"B) {B_key}", variable=radio, value=2,fg="#4572F9", command = sellect)  
            R2.place(x = 20, y=190, height=50, width=860, anchor = "w")  
        
            R3 = Radiobutton(frame_4_3, text=f"C) {C_key}", variable=radio, value=3,fg="#4572F9", command = sellect)  
            R3.place(x = 20, y=260, height=50, width=860,anchor = "w")

            R4 = Radiobutton(frame_4_3, text=f"D) {D_key}", variable=radio, value=4,fg="#4572F9", command = sellect)  
            R4.place(x = 20, y=330, height=50, width=860, anchor = "w")
    
        #l3_1 = Label(frame_4_3, text="Ushbu testda har bir savol uchun 1,5 daqiqa beriladi.", font=("Arial", 15), bg="red", fg='white')
        #l3_1.place(x = 150, y = 200, height=30, width=600)

        btn_7 = Button(frame_4_3, text = "ortga", bd = 6, activebackground = '#1C4BD6', activeforeground = "white", fg='white', bg = '#001659', command = ortga)
        btn_7.place(x = 700, y = 450, height = 30, width = 80)

        btn_8 = Button(frame_4_3, text = "oldinga", bd = 6, activebackground = '#1C4BD6', activeforeground = "white", fg='white', bg = '#001659', command = oldinga )
        btn_8.place(x = 800, y = 450, height = 30, width = 80)
        oldinga()
    except:
        showinfo(
            title="File Error",
            message = '''File Tanlamadiz!!!!'''

            )
        boshlash()
    
def add_help():
    showinfo(
        title='Add File Help',
        message='''File qoshish uchun faqat xlsx filelar qabul qiladi.
birinchi savollar qolgan ustunlarda '''
    )
    

def boshlash():
    frame_1.destroy()
    
    frame_6 = Frame(root, bg="#4572F9")
    frame_6.place(x = 0, y = 0, height=50, width=900)   
    
    

    l2 = Label(frame_6, text = f"{baza[m][0]} {baza[m][1]}", font=("Arial", 15), bg="#4572F9", fg='white')
    l2.place(x = 300, y = 15, height=30, width=300)

    #l3 = Label(frame_6, text=f"Umumiy vaqt: {minut} : {s}", font=("Arial", 15), bg="#4572F9", fg='white')
    #l3.place(x = 600, y = 15, height=30, width=300)

    frame_7 = Frame(root, bg="#3B4B7D")
    frame_7.place(x = 0, y = 550, height=50, width=900)

    
    frame_8 = Frame(root, bg="#1B2A58")
    frame_8.place(x = 0, y = 50, height=500, width=900)

    btn_7 = Button(frame_8, text = "Start", bd = 6, activebackground = '#1C4BD6', activeforeground = "white", fg='white', bg = '#001659' ,command = test)
    btn_7.place(x = 340, y = 200, height = 50, width = 100)

    btn_8 = Button(frame_8, text = "Open to file", bd = 6, activebackground = '#1C4BD6', activeforeground = "white", fg='white', bg = '#001659',command=select_file )
    btn_8.place(x = 460, y = 200, height = 50, width = 100)

    btn_9 = Button(frame_8, text = "Add File Help", bd = 6, activebackground = '#1C4BD6', activeforeground = "white", fg='white', bg = '#001659',command=add_help )
    btn_9.place(x = 340, y = 270, height = 50, width = 220)
        
    
main() 
root.mainloop()
